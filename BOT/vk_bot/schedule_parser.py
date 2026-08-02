"""
Парсер расписания из Excel-файлов ИрГАУ.
Поддерживает все направления: 09.03.03, 38.05.01, 38.03.02, 38.03.01.

Структура Excel:
  строка 2  — название направления
  строка 4  — номер курса
  строка 5  — номера групп (col[3]=1, col[5]=2 если есть 2 группы)
  строки 6+ — расписание:
    col[1] — день недели (ПОНЕДЕЛЬНИК …) — может быть в той же строке, что и пара!
    col[2] — время пары
    col[3] — занятие гр.1 / общая лекция
    col[4] — ауд. гр.1 (None при общей лекции)
    col[5] — занятие гр.2 (None при общей лекции / у одногрупповых файлов)
    col[6] — ауд. гр.2 / общая ауд.
"""
import os
import re
import openpyxl
from datetime import datetime

SCHEDULES_DIR = os.path.join(os.path.dirname(__file__), "schedules")

DIRECTIONS: dict[str, dict] = {
    "09.03.03": {"name": "Прикладная информатика",    "short": "ПИ"},
    "38.05.01": {"name": "Экономическая безопасность","short": "ЭБ"},
    "38.03.02": {"name": "Менеджмент",                "short": "МЕН"},
    "38.03.01": {"name": "Экономика",                 "short": "ЭК"},
}

DAYS_RU = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА"]
WEEKDAY_TO_DAY = {0:"ПОНЕДЕЛЬНИК",1:"ВТОРНИК",2:"СРЕДА",3:"ЧЕТВЕРГ",4:"ПЯТНИЦА",5:"СУББОТА",6:"ВОСКРЕСЕНЬЕ"}
DAY_ALIASES: dict[str, str] = {
    "понедельник":"ПОНЕДЕЛЬНИК","пн":"ПОНЕДЕЛЬНИК",
    "вторник":"ВТОРНИК","вт":"ВТОРНИК",
    "среда":"СРЕДА","ср":"СРЕДА","среду":"СРЕДА",
    "четверг":"ЧЕТВЕРГ","чт":"ЧЕТВЕРГ",
    "пятница":"ПЯТНИЦА","пт":"ПЯТНИЦА","пятницу":"ПЯТНИЦА",
    "суббота":"СУББОТА","сб":"СУББОТА","субботу":"СУББОТА",
}


# ─── Вспомогательные ──────────────────────────────────────────────────────────

def _file(direction: str) -> str:
    return os.path.join(SCHEDULES_DIR, f"{direction}.xlsx")


def _sheets(direction: str) -> list[str]:
    wb = openpyxl.load_workbook(_file(direction), data_only=True, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _detect_groups(ws) -> int:
    """Читает строку 5 (индекс 4) и определяет количество групп."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 4:  # строка 5 (0-based)
            # col[5] содержит «2» — вторая группа
            try:
                val = row[5]
                if val is not None and str(val).strip() == "2":
                    return 2
            except IndexError:
                pass
            return 1
    return 1


def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _parse_sheet(ws, group: int, num_groups: int) -> dict[str, list[dict]]:
    """
    Возвращает {день: [{"time": str, "lesson": str, "room": str}]}.
    ИСПРАВЛЕНО: строка с названием дня недели может одновременно содержать пару.
    """
    if num_groups == 2:
        # 9-колоночный формат (09.03.03)
        l_col  = 3 if group == 1 else 5   # колонка занятия для своей группы
        r_col  = 4 if group == 1 else 6   # колонка ауд. для своей группы
        l_shared = 3                        # общая лекция всегда в col[3]
        r_shared = 6                        # общая ауд. всегда в col[6]
    else:
        # 7-колоночный формат (одна группа)
        l_col = l_shared = 3
        r_col = r_shared = 4

    result: dict[str, list[dict]] = {}
    current_day: str | None = None

    for row in ws.iter_rows(values_only=True):
        day_cell  = _str(row[1]) if len(row) > 1 else ""
        time_cell = _str(row[2]) if len(row) > 2 else ""

        # ── Обновляем текущий день (НЕ делаем continue — та же строка может
        #    содержать пару, например первая пара дня в 8:30) ─────────────────
        if day_cell.upper() in DAYS_RU:
            current_day = day_cell.upper()
            result.setdefault(current_day, [])

        if not current_day or not time_cell:
            continue
        if not any(c.isdigit() for c in time_cell):
            continue

        # ── Определяем занятие ───────────────────────────────────────────────
        lesson = ""
        room   = ""

        if num_groups == 2 and group == 2:
            # Для группы 2: сначала ищем в своей колонке, потом общую лекцию
            l2 = _str(row[l_col])  if len(row) > l_col  else ""
            r2 = _str(row[r_col])  if len(row) > r_col  else ""
            if l2:
                lesson, room = l2, r2 or (_str(row[r_shared]) if len(row) > r_shared else "")
            else:
                # Возможно, это общая лекция
                ls = _str(row[l_shared]) if len(row) > l_shared else ""
                rs = _str(row[r_shared]) if len(row) > r_shared else ""
                if ls:
                    lesson, room = ls, rs
        else:
            lesson = _str(row[l_col]) if len(row) > l_col else ""
            room   = _str(row[r_col]) if len(row) > r_col else ""
            # Если ауд. пустая — попробовать r_shared
            if lesson and not room and r_shared != r_col:
                room = _str(row[r_shared]) if len(row) > r_shared else ""

        if not lesson:
            continue

        # Очищаем лишние пробелы в аудитории
        room = re.sub(r"\s{2,}", "  ", room).strip()

        result[current_day].append({
            "time":   time_cell,
            "lesson": lesson,
            "room":   room,
        })

    return result


def _load(direction: str, course: int, group: int) -> dict[str, list[dict]]:
    wb  = openpyxl.load_workbook(_file(direction), data_only=True)
    sheet_name = f"{course} курс"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Курс {course} не найден в расписании {direction}")
    ws = wb[sheet_name]
    ng = _detect_groups(ws)
    if ng == 1 and group != 1:
        group = 1
    data = _parse_sheet(ws, group, ng)
    wb.close()
    return data


def _format_day(day_name: str, lessons: list[dict]) -> str:
    cap = day_name.capitalize()
    if not lessons:
        return f"📅 {cap}: выходной / нет занятий"
    lines = [f"📅 {cap}:"]
    for p in lessons:
        line = f"  🕐 {p['time']}  {p['lesson']}"
        if p["room"]:
            line += f"  (ауд. {p['room']})"
        lines.append(line)
    return "\n".join(lines)


def _header(direction: str, course: int, group: int, num_groups: int) -> str:
    d = DIRECTIONS.get(direction, {})
    name = d.get("name", direction)
    g_str = f", группа {group}" if num_groups == 2 else ""
    return f"📚 {direction} — {name}\n{course} курс{g_str}\n"


# ─── Публичный API ────────────────────────────────────────────────────────────

def list_directions() -> str:
    lines = ["📋 Доступные направления:\n"]
    for code, info in DIRECTIONS.items():
        sheets = _sheets(code)
        courses = len([s for s in sheets if "курс" in s])
        lines.append(f"  {code} — {info['name']} ({courses} курса/ов)")
    lines.append(
        "\nЧтобы выбрать, напиши, например:\n"
        "  направление 09.03.03\n"
        "  или: направление менеджмент"
    )
    return "\n".join(lines)


def get_num_groups(direction: str, course: int) -> int:
    wb = openpyxl.load_workbook(_file(direction), data_only=True)
    sheet_name = f"{course} курс"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 1
    ng = _detect_groups(wb[sheet_name])
    wb.close()
    return ng


def get_schedule(direction: str, course: int, group: int, mode: str = "week") -> str:
    """mode: 'week' | 'today' | 'tomorrow' | day name (ПОНЕДЕЛЬНИК etc.)"""
    try:
        data = _load(direction, course, group)
    except (ValueError, FileNotFoundError) as e:
        return f"❌ {e}"

    ng = get_num_groups(direction, course)
    hdr = _header(direction, course, group, ng)

    now = datetime.now()
    today    = WEEKDAY_TO_DAY[now.weekday()]
    tomorrow = WEEKDAY_TO_DAY[(now.weekday() + 1) % 7]

    if mode == "today":
        return hdr + "\n" + _format_day(today, data.get(today, []))
    if mode == "tomorrow":
        return hdr + "\n" + _format_day(tomorrow, data.get(tomorrow, []))
    if mode.upper() in DAYS_RU:
        d = mode.upper()
        return hdr + "\n" + _format_day(d, data.get(d, []))

    # Вся неделя
    parts = [hdr]
    for day in DAYS_RU:
        lessons = data.get(day, [])
        if lessons:
            parts.append(_format_day(day, lessons))
    if len(parts) == 1:
        parts.append("На эту неделю занятий не найдено.")
    return "\n\n".join(parts)


def get_free_windows(direction: str, course: int, group: int, mode: str = "today") -> str:
    """Ищет «окна» — промежутки между парами."""
    try:
        data = _load(direction, course, group)
    except (ValueError, FileNotFoundError) as e:
        return f"❌ {e}"

    ng = get_num_groups(direction, course)
    hdr = _header(direction, course, group, ng)

    now = datetime.now()
    if mode == "today":
        days_to_check = [WEEKDAY_TO_DAY[now.weekday()]]
        title = "Окна на сегодня"
    elif mode == "tomorrow":
        days_to_check = [WEEKDAY_TO_DAY[(now.weekday() + 1) % 7]]
        title = "Окна на завтра"
    else:
        days_to_check = DAYS_RU
        title = "Свободные окна на неделю"

    parts = [hdr + f"\n🪟 {title}:\n"]
    found_any = False

    for day in days_to_check:
        lessons = data.get(day, [])
        if len(lessons) < 2:
            continue

        windows = []
        for i in range(len(lessons) - 1):
            end_cur   = lessons[i]["time"].split("-")[1].strip()
            start_nxt = lessons[i + 1]["time"].split("-")[0].strip()
            if end_cur != start_nxt:
                windows.append(f"  🕐 {end_cur} — {start_nxt}  (между {lessons[i]['time']} и {lessons[i+1]['time']})")

        if windows:
            found_any = True
            parts.append(f"📅 {day.capitalize()}:")
            parts.extend(windows)

    if not found_any:
        parts.append("Свободных окон нет — пары идут подряд.")
    return "\n".join(parts)


def search_by_teacher(surname: str) -> str:
    """Ищет все пары с преподавателем по фамилии во всех направлениях."""
    surname_lo = surname.strip().lower()
    results: list[str] = []

    for direction in DIRECTIONS:
        try:
            sheets = _sheets(direction)
        except Exception:
            continue
        for sheet_name in sheets:
            if "курс" not in sheet_name:
                continue
            try:
                wb = openpyxl.load_workbook(_file(direction), data_only=True)
                ws = wb[sheet_name]
                ng = _detect_groups(ws)
                for g in range(1, ng + 1):
                    data = _parse_sheet(ws, g, ng)
                    for day, lessons in data.items():
                        for p in lessons:
                            if surname_lo in p["lesson"].lower():
                                g_str = f" гр.{g}" if ng == 2 else ""
                                room = f" (ауд. {p['room']})" if p["room"] else ""
                                results.append(
                                    f"  {direction} {sheet_name}{g_str} | "
                                    f"{day.capitalize()} {p['time']}{room}\n"
                                    f"  {p['lesson']}"
                                )
                wb.close()
            except Exception:
                continue

    if not results:
        return f"🔍 Преподаватель «{surname}» не найден ни в одном расписании."

    header = f"🔍 Пары с преподавателем «{surname.capitalize()}»:\n"
    # Ограничиваем вывод
    if len(results) > 20:
        results = results[:20]
        results.append("  … (показаны первые 20 результатов)")
    return header + "\n\n".join(results)
